"""
Definition of views.
"""

from django.views.generic.detail import DetailView
from django.views.generic.list import ListView
from django.shortcuts import get_object_or_404, render
from django.http import HttpRequest, HttpResponse
from django.core.urlresolvers import reverse
from django.template import RequestContext, loader
from django.template.loader import render_to_string
from django.db import connection
from django.db.models import Q
from django.db.models.functions import Lower
from django.http import JsonResponse
from datetime import datetime
from xml.dom import minidom
from operator import itemgetter
import xml.etree.ElementTree as ET
import os
import operator
import re
import fnmatch
import csv
import codecs
import copy
from wld.dictionary.models import *
from wld.dictionary.forms import *
#from wld.dictionary.adminviews import order_queryset_by_sort_order
from wld.settings import APP_PREFIX, WSGI_FILE

# Global variables
paginateSize = 10
paginateLocations = 2
paginateEntries = 100
# paginateValues = (1000, 500, 250, 100, 50, 40, 30, 20, 10, )
paginateValues = (100, 50, 20, 10, 5, 2, 1, )
outputColumns = ['begrip', 'trefwoord', 'dialectopgave', 'Kloekecode', 'aflevering', 'bronnenlijst']

THIS_DICTIONARY = "e-WLD"

# General help functions
def order_queryset_by_sort_order(get, qs, sOrder = 'gloss'):
    """Change the sort-order of the query set, depending on the form field [sortOrder]

    This function is used by EntryListView.
    The value of [sortOrder] is 'woord' by default.
    [sortOrder] is a hidden field inside the "adminsearch" html form in the template admin_gloss_list.html
    Its value is changed by clicking the up/down buttons in the second row of the search result table
    """

    def get_string_from_tuple_list(lstTuples, number):
        """Get the string value corresponding to a number in a list of number-string tuples"""
        sBack = [tup[1] for tup in lstTuples if tup[0] == number]
        return sBack

    # Helper: order a queryset on field [sOrder], which is a number from a list of tuples named [sListName]
    def order_queryset_by_tuple_list(qs, sOrder, sListName):
        """Order a queryset on field [sOrder], which is a number from a list of tuples named [sListName]"""

        # Get a list of tuples for this sort-order
        tpList = build_choice_list(sListName)
        # Determine sort order: ascending is default
        bReversed = False
        if (sOrder[0:1] == '-'):
            # A starting '-' sign means: descending order
            sOrder = sOrder[1:]
            bReversed = True

        # Order the list of tuples alphabetically
        # (NOTE: they are alphabetical from 'build_choice_list()', except for the values 0,1)
        tpList = sorted(tpList, key=operator.itemgetter(1))
        # Order by the string-values in the tuple list
        return sorted(qs, key=lambda x: get_string_from_tuple_list(tpList, getattr(x, sOrder)), reverse=bReversed)

    # Set the default sort order
    # sOrder = 'gloss'  # Default sort order if nothing is specified
    # See if the form contains any sort-order information
    if ('sortOrder' in get and get['sortOrder'] != ''):
        # Take the user-indicated sort order
        sOrder = get['sortOrder']

    # The ordering method depends on the kind of field:
    # (1) text fields are ordered straightforwardly
    # (2) fields made from a choice_list need special treatment
    if (sOrder.endswith('handedness')):
        ordered = order_queryset_by_tuple_list(qs, sOrder, "Handedness")
    else:
        # Use straightforward ordering on field [sOrder]
        ordered = qs.order_by(Lower(sOrder))

    # return the ordered list
    return ordered

def get_item_list(lVar, lFun, qs):
    """Turn the queryset [qs] into a list of Items that have first and last information"""

    lItem = []
    oErr = ErrHandle()
    try:
        # Initialize the variables whose changes are important
        oVariable = {}
        for i, key in enumerate(lVar):
            oVariable[key] = "" # {'name': key, 'fun': lFun[i]}
        iLast = len(qs)-1
        # Iterate over the entries looking for first, last etc
        for i, entry in enumerate(qs):
            bIsLastEntry = (i==iLast)
            oItem = {'entry': entry}
            for k in lVar:
                oItem[k] = {'first':False, 'last':False}
            bIsDict = isinstance(entry, dict)
            bVarIsLast = False
            # Check for changes in all the variables
            for j, k in enumerate(lVar):
                fun = lFun[j]
                if callable(fun):
                    sValue = fun(entry)
                else:
                    for idx, val in enumerate(fun):
                        if idx==0:
                            if bIsDict:
                                sValue = entry[val]
                            else:
                                sValue = getattr(entry, val)
                        else:
                            if bIsDict:
                                sValue = sValue[val]
                            else:
                                sValue = getattr(sValue, val)
                # Check for changes in the value of the variable 
                # if sValue != oVariable[k]:
                if sValue != oVariable[k] or bVarIsLast or (i>0 and lItem[i-1][k]['last']):
                    # Check if the previous one's [last] must be changed
                    if oVariable[k] != "": lItem[i-1][k]['last'] = True
                    # Adapt the current one's [first] property
                    oItem[k]['first']= True
                    # Adapt the variable
                    oVariable[k] = sValue      
                    # Indicate that the next ones should be regarded as 'last'
                    bVarIsLast = True      
                # Check if this is the last
                if bIsLastEntry: oItem[k]['last'] = True
            # Add this object to the list of items
            lItem.append(oItem)
    except:
        oErr.DoError("get_item_list error")
        lItem = []

    # Return the list we have made
    return lItem

def home(request):
    """Renders the home page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'dictionary/index.html',
        {
            'title': THIS_DICTIONARY,
            'year':datetime.now().year,
        }
    )

def contact(request):
    """Renders the contact page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'dictionary/contact.html',
        {
            'title':'{} contact'.format(THIS_DICTIONARY),
            'message':'Henk van den Heuvel (H.vandenHeuvel@Let.ru.nl)',
            'year':datetime.now().year,
        }
    )

def about(request):
    """Renders the about page."""
    assert isinstance(request, HttpRequest)
    return render(request, 'dictionary/about.html',
        {   'title':'{} informatie'.format(THIS_DICTIONARY),
            'message':'Radboud Universiteit Nijmegen - Dialectenwoordenboek.',
            'year':datetime.now().year,
        }
    )

def guide(request):
    """Renders the 'guide' page."""
    assert isinstance(request, HttpRequest)
    return render( request, 'dictionary/guide.html',
        {   'title':'{} handleiding'.format(THIS_DICTIONARY),
            'message':'Radboud Universiteit Nijmegen - Dialectenwoordenboek.',
            'dic_abbr': THIS_DICTIONARY,
            'year':datetime.now().year,
        }
    )

def do_repair(request):
    """Renders the REPAIR page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'dictionary/repair.html',
        {   'title':'{} reparatie'.format(THIS_DICTIONARY),
            'message':'Radboud Universiteit Nijmegen - Dialectenwoordenboek.',
            'year':datetime.now().year,
        }
    )

def adapt_search(val):
    # First trim
    val = val.strip()
    # fnmatch.translate() works okay, but note beginning and ending spaces
    # val = fnmatch.translate('^' + val + '$')
    val = '^' + fnmatch.translate(val) + '$'
    # val = '^' + val.replace("?", ".").replace("*", ".*") + '$'
    return val

def export_csv(qs, sFileName):
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="'+sFileName+'.csv"'
     
    # Create a writer for the CSV
    # writer = csv.writer(response, csv.excel, delimiter='\t')
    writer = csv.writer(response, csv.excel_tab)
    # BOM to indicate that this is UTF8
    response.write(u'\ufeff'.encode('utf8'))
    # Output the first row with the headings -- these take 'entry' as basis
    fields = outputColumns
    # Output the first row with the headings
    writer.writerow(fields)
    # Walk through the queryset
    for obj in qs:
        writer.writerow(obj.get_row())

    return response

def export_xlsx(qs, sFileName):
    import openpyxl
    from openpyxl.utils.cell import get_column_letter
    from openpyxl import Workbook

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename='+sFileName+'.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = sFileName

    row_num = 0

    columns = outputColumns
    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num]
        # c.style.font.bold = True
        c.font = openpyxl.styles.Font(bold=True)
        # Set column width
        # ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]
        # ws.column_dimensions[get_column_letter(col_num+1)].bestFit = True
        ws.column_dimensions[get_column_letter(col_num+1)].width = 20.0

    # Walk the queryset
    for obj in qs:
        row_num += 1
        row = obj.get_row()
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.alignment = openpyxl.styles.Alignment(wrap_text=True)

    wb.save(response)
    return response

def export_html(qs, sFileName):
    # Create a string with the HTML contents
    arOut = []
    arOut.append("<html><head><meta charset='utf-8' /></head><body><table><thead><tr>")
    fields = outputColumns
    for f in fields:
        arOut.append("<th>"+f+"</th>")
    arOut.append("</tr></thead>")
    arOut.append("<tbody>")
    # Walk the contents
    for obj in qs:
        sRow = "<tr><td>"+obj.lemma.gloss+"</td><td>"+obj.trefwoord.woord+"</td><td>"+obj.woord+"</td><td>"+obj.dialect.nieuw+"</td><td>"+obj.aflevering.naam+"</td><tr>"
        arOut.append(sRow)
    content = "\n".join(arOut)
    # Create the HttpResponse object with the appropriate header.
    response = HttpResponse(content, content_type='text/html')
    response['Content-Disposition'] = 'attachment; filename="'+sFileName+'.htm"'
     
    return response

def do_repair_start(request):
    """Start up the repair action"""
    sRepairType = request.GET.get('repairtype', '')

    # Formulate a response
    data = {'status': 'done'}

    # Remove any previous repair objects of this type
    Repair.objects.filter(repairtype=sRepairType).delete()
    # Retrieve the Repair object with the correct type
    oRepair = Repair(repairtype=sRepairType)
    oRepair.save()
    if sRepairType == "lemma":
        bResult = do_repair_lemma(oRepair)
        if not bResult:
            data.status = "error"
    elif sRepairType == "entrydescr":
        bResult = do_repair_entrydescr(oRepair)
        if not bResult:
            data.status = "error"

    # Return this response
    return JsonResponse(data)

def do_repair_progress(request):
    """Show the progress of where we are in repairing"""

    sRepairType = request.GET.get('repairtype', '')

    # Formulate a response
    data = {'status': 'not found'}

    # Get the repair object
    qs = Repair.objects.filter(repairtype=sRepairType)
    if qs != None and len(qs) > 0:
        oRepair = qs[0]
        data['status'] = oRepair.status

    # Return this response
    return JsonResponse(data)


def import_csv_start(request):
    # Formulate a response
    data = {'status': 'done'}
    oErr = ErrHandle()

    try:
        # x = request.POST
        iDeel = request.GET.get('deel', 1)
        iSectie = request.GET.get('sectie', None)
        iAflnum = request.GET.get('aflnum', 1)
        sFile = request.GET.get('filename', '')

        bUseDbase = request.GET.get('usedbase', False)
        if bUseDbase:
            if bUseDbase == "true":
                bUseDbase = True
            else:
                bUseDbase = False

        # Get the id of the Info object
        if iSectie==None or iSectie == "":
            info = Info.objects.filter(deel=iDeel, aflnum=iAflnum).first()
        else:
            info = Info.objects.filter(deel=iDeel, sectie=iSectie, aflnum=iAflnum).first()

        if info == None:
            data['status'] = 'error: no Info object found'
            return JsonResponse(data)

        # Remove any previous status objects for this info
        Status.objects.filter(info=info).delete()

        # Create a new import-status object
        oStatus = Status(info=info)

        # Note that we are starting
        oStatus.set_status("starting")
        iStatus = oStatus.id
        # oCsvImport['status'] = "starting"

        # Call the process
        oResult = csv_to_fixture(sFile, iDeel, iSectie, iAflnum, iStatus, bUseDbase = bUseDbase, bUseOld = True)
        if oResult == None or oResult['result'] == False:
            data['status'] = 'error'

        # WSince we are done: explicitly set the status so
        oStatus.set_status("done")
    except Exception as ex:
        oErr.DoError("import_csv_start error")
        data['status'] = "error"

    # Return this response
    return JsonResponse(data)

def import_csv_progress(request):
    oErr = ErrHandle()
    # Prepare a return object
    data = {'read':0, 'skipped':0, 'method': '(unknown)', 'msg': ''}
    try:
        # Debugging
        oErr.Status("import_csv_progress at {}".format(datetime.now()))

        if request.POST:
            qd = request.POST
        else:
            qd = request.GET
        iDeel = qd.get('deel', 1)
        iSectie = qd.get('sectie', None)
        iAflnum = qd.get('aflnum', 1)
        # Get the id of the Info object
        if iSectie==None or iSectie == "":
            info = Info.objects.filter(deel=iDeel, aflnum=iAflnum).first()
        else:
            info = Info.objects.filter(deel=iDeel, sectie=iSectie, aflnum=iAflnum).first()
        # Find out how far importing is going
        qs = Status.objects.filter(info=info)
        if qs != None and len(qs) > 0:
            oStatus = qs[0]
            # Fill in the return object
            data['read'] = oStatus.read
            data['skipped'] = oStatus.skipped
            data['method'] = oStatus.method
            data['status'] = oStatus.status
            # Checking...
            if data['status'] == "idle":
                data['msg'] = "Idle status in import_csv_progress"
        else:
            # Do we have an INFO object?
            if info == None:
                data['status'] = "Please supply [deel], [sectie] and [aflnum]"
            else:
                data['status'] = "No status object for info=" + str(info.id) + " has been created yet"
    except Exception as ex:
        oErr.DoError("import_csv_progress error")
        data['status'] = "error"

    # Return where we are
    return JsonResponse(data)


class DictionaryDetailView(DetailView):
    """Details of an entry from the dictionary"""

    model = Entry
    context_object_name = 'entry'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(DictionaryDetailView, self).get_context_data(**kwargs)

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Set the title of the application
        context['title'] = "{} detail".format(THIS_DICTIONARY)

        # Return the calculated context
        return context


class TrefwoordListView(ListView):
    """ListView of keywords (trefwoorden)"""

    model = Trefwoord
    template_name = 'dictionary/trefwoord_list.html'
    # paginate_by = paginateEntries
    paginate_by = paginateSize
    entrycount = 0
    bUseMijnen = True   # Limburg uses mijnen, Brabant not
    bWbdApproach = True # Filter using the WBD approach (this also applies for the revised WLD)
    qEntry = None
    qs = None
    bDoTime = True      # Measure time
    strict = True      # Use strict filtering

    def get_qs(self):
        if self.qEntry == None:
            if self.qs != None:
                qs = self.qs
            else:
                # Get the PKs of Entry related to Trefwoord
                qs = self.get_queryset()
            # Get the Entry queryset related to this
            qs = Entry.objects.filter(trefwoord__pk__in=qs).select_related()
        else:
            qs = self.qEntry
        return qs

    def render_to_response(self, context, **response_kwargs):
        """Check if a CSV response is needed or not"""
        if 'Csv' in self.request.GET.get('submit_type', ''):
            """ Provide CSV response"""
            return export_csv(self.get_qs(), 'trefwoorden')
        elif 'Excel' in self.request.GET.get('submit_type', ''):
            """ Provide Excel response"""
            return export_xlsx(self.get_qs(), 'trefwoorden')
        elif 'Html' in self.request.GET.get('submit_type', ''):
            """ Provide Html response"""
            return export_html(self.get_qs(), 'trefwoorden')
        else:
            oResponse = super(TrefwoordListView, self).render_to_response(context, **response_kwargs)
            return oResponse

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(TrefwoordListView, self).get_context_data(**kwargs)

        # Action depends on the approach
        if self.bWbdApproach:
            # Need to adapt the object_list to get the entries to be used
            context['object_list'] = self.get_entryset(context['page_obj'])

        # Get parameters for the search
        initial = self.request.GET
        if initial is None:
            initial = {'optdialect': 'stad'}

        # Fill the 'searchform' context variable with the values that are received from the GET request
        search_form = TrefwoordSearchForm(initial)
        context['searchform'] = search_form

        if 'paginate_by' in initial:
            context['paginateSize'] = int(initial['paginate_by'])
            self.paginate_by = int(initial['paginate_by'])
        else:
            context['paginateSize'] = self.paginate_by  # paginateSize

        if self.bUseMijnen:
            # Try to retain the choice for Mijn
            if 'mijn' in initial:
                mijn_id = int(initial['mijn'])
                context['mijnkeuze'] = mijn_id
                mijn_inst = Mijn.objects.filter(id=mijn_id).first()
                if mijn_inst == None:
                    context['mijnnaam'] = ''
                else:
                    context['mijnnaam'] = mijn_inst.naam
            else:
                context['mijnkeuze'] = 0
                context['mijnnaam'] = ''
        # Process and retain the choice for Aflevering
        if 'aflevering' in initial:
            context['aflkeuze'] = int(initial['aflevering'])
            afl = Aflevering.objects.filter(id=context['aflkeuze']).first()
            if afl == None:
                context['afl'] = ''
            else:
                context['afl'] = afl.get_summary()
        else:
            context['aflkeuze'] = 0
            context['afl'] = ''

        # Get possible user choice of 'strict'
        if 'strict' in initial:
            self.strict = (initial['strict'] == "True")
        context['strict'] = str(self.strict)
        
        # Determine the count 
        context['entrycount'] = self.entrycount   #  self.get_queryset().count()
        # context['twcount'] = self.twcount

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Set the afleveringen and mijnen that are available
        context['afleveringen'] = [afl for afl in Aflevering.objects.all()]
        context['mijnen'] = [mijn for mijn in Mijn.objects.all().order_by('naam')]

        # Set the title of the application
        context['title'] = "{} trefwoorden".format(THIS_DICTIONARY)

        # If we are in 'strict' mode, we need to deliver the [qlist]
        if self.strict:
            # Transform the paginated queryset into a dict sorted by Dialect/Aflevering
            lAflev = self.get_qafl(context)

            # Get a list with 'first' and 'last' values for each item in the current paginated queryset
            lEntry = self.get_qlist(context)
            # Add the sorted-dialect information to lEntry
            for idx, item in enumerate(lEntry):
                # Start or Finish dialect information
                if item['trefwoord_woord']['first']:
                    qsa = []
                # All: add this entry
                qsa.append(lAflev[idx])
                if item['trefwoord_woord']['last']:
                    # COpy the list of Entry elements sorted by Trefwoord/Aflevering here
                    lEntry[idx]['alist'] = qsa
                else:
                    lEntry[idx]['alist'] = None

            context['qlist'] = lEntry

        # Return the calculated context
        return context
      
    def get_qlist(self, context):
        """Get a list of Entry elements + first/last information"""

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']
        # Start the output
        html = []
        # Initialize the variables whose changes are important
        lVars = ["trefwoord_woord", "lemma_gloss", "toelichting", "dialectopgave", "dialect_stad"]
        lFuns = [["trefwoord", "woord"], ["lemma", "gloss"], Entry.get_toelichting, Entry.dialectopgave, ["dialect", "stad"]]
        # Get a list of items containing 'first' and 'last' information
        lItem = get_item_list(lVars, lFuns, qs)
        # REturn this list
        return lItem

    def get_qafl(self, context):
        """Sort the paginated QS by Trefwoord/Aflevering into a list"""

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']
        qsd = []
        # Walk through the query set
        for entry in qs:
            qsd.append(entry)
        # Now sort the resulting set
        qsd = sorted(qsd, key=lambda el: el.trefwoord.woord + " " + el.get_aflevering())
        # Prepare for processing
        lVarsD = ["trefw", "afl"]
        lFunsD = [["trefwoord", "woord"], Entry.get_aflevering]
        # Create a list of the items
        lAfl = get_item_list(lVarsD, lFunsD, qsd)
        # Return the result
        return lAfl            
      
    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('paginate_by', self.paginate_by)
        
    def get_entryset(self, page_obj):
        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Retrieve the set of trefwoorden from the page_obj
        trefw_list = [item.id for item in page_obj]

        # Initialize the filtering
        lstQ.append(Q(trefwoord__id__in=trefw_list))

        # Get the parameters passed on with the GET request
        get = self.request.GET

        # Check for aflevering being publishable
        if self.strict:
            lstQ.append(Q(aflevering__toonbaar=True))
        else:
            lstQ.append(Q(entry__aflevering__toonbaar=True))

        # Check for dialectwoord
        if 'dialectwoord' in get and get['dialectwoord'] != '':
            val = adapt_search(get['dialectwoord'])
            # Adapt Entry filter
            if self.strict:
                lstQ.append(Q(woord__iregex=val))
            else:
                lstQ.append(Q(entry__woord__iregex=val))
            bHasFilter = True

        # Check for lemma
        if 'lemma' in get and get['lemma'] != '':
            val = adapt_search(get['lemma'])
            # Adapt Entry filter
            if self.strict:
                lstQ.append(Q(lemma__gloss__iregex=val))
            else:
                lstQ.append(Q(entry__lemma__gloss__iregex=val))
            bHasFilter = True

        # Check for dialect city
        if 'dialectCity' in get and get['dialectCity'] != '':
            val = adapt_search(get['dialectCity'])
            # Adapt Entry filter
            if self.strict:
                lstQ.append(Q(dialect__stad__iregex=val))
            else:
                lstQ.append(Q(entry__dialect__stad__iregex=val))
            bHasFilter = True

        # Check for dialect code (Kloeke)
        if 'dialectCode' in get and get['dialectCode'] != '':
            val = adapt_search(get['dialectCode'])
            # Adapt Entry filter
            if self.strict:
                lstQ.append(Q(dialect__nieuw__iregex=val))
            else:
                lstQ.append(Q(entry__dialect__nieuw__iregex=val))
            bHasFilter = True

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    if self.strict:
                        lstQ.append(Q(aflevering__id=iVal))
                    else:
                        lstQ.append(Q(entry__aflevering__id=iVal))
                    bHasFilter = True

        # Check for mijn
        if self.bUseMijnen and 'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    if self.strict:
                        lstQ.append(Q(mijnlijst__id=iVal))
                    else:
                        lstQ.append(Q(entry__mijnlijst__id=iVal))
                    bHasFilter = True

        # Order: "trefwoord_woord", "lemma_gloss", "dialectopgave", "dialect_stad"
        # Make sure we apply the filter
        qse = Entry.objects.filter(*lstQ).distinct().select_related().order_by(
            Lower('trefwoord__woord'), 
            'lemma__gloss',  
            Lower('toelichting'),
            Lower('woord'), 
            Lower('dialect__stad'))
        self.qEntry = qse
        return qse
        
    def get_queryset(self):

        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        
        # Debugging: mesaure time
        if self.bDoTime: iStart = get_now_time()

        # Get possible user choice of 'strict'
        if 'strict' in get:
            self.strict = (get['strict'] == "True")

        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Fine-tuning: search string is the LEMMA
        if 'search' in get and get['search'] != '':
            val = get['search']
            if '*' in val or '[' in val or '?' in val:
                val = adapt_search(val)
                lstQ.append(Q(woord__iregex=val) )
            else:
                # Strive for equality, but disregard case
                lstQ.append(Q(woord__iexact=val))
            #val = adapt_search(get['search'])
            ## Use the 'woord' attribute of Trefwoord
            #lstQ.append(Q(woord__iregex=val) )
            bHasSearch = True

            # check for possible exact numbers having been given
            if re.match('^\d+$', val):
                lstQ.append(Q(sn__exact=val))

        # Check for 'toelichting'
        if 'toelichting' in get and get['toelichting'] != '':
            val = adapt_search(get['toelichting'])
            # Try to get to the 'toelichting'
            lstQ.append(Q(toelichting__iregex=val))
            bHasSearch = True

        # Check for dialectwoord
        if 'dialectwoord' in get and get['dialectwoord'] != '':
            val = adapt_search(get['dialectwoord'])
            # Adapt Entry filter
            lstQ.append(Q(entry__woord__iregex=val))
            bHasFilter = True

        # Check for lemma
        if 'lemma' in get and get['lemma'] != '':
            val = adapt_search(get['lemma'])
            # Adapt Entry filter
            lstQ.append(Q(entry__lemma__gloss__iregex=val))
            bHasFilter = True

        # Check for dialect city
        if 'dialectCity' in get and get['dialectCity'] != '':
            val = adapt_search(get['dialectCity'])
            # Adapt Entry filter
            lstQ.append(Q(entry__dialect__stad__iregex=val))
            bHasFilter = True

        # Check for dialect code (Kloeke)
        if 'dialectCode' in get and get['dialectCode'] != '':
            val = adapt_search(get['dialectCode'])
            # Adapt Entry filter
            lstQ.append(Q(entry__dialect__nieuw__iregex=val))
            bHasFilter = True

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__aflevering__id=iVal))
                    bHasFilter = True

        # Check for mijn
        if self.bUseMijnen and 'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__mijnlijst__id=iVal))
                    bHasFilter = True

        # Debugging: time
        if self.bDoTime: print("TrefwoordListView get_queryset part 1: {:.1f}".format(get_now_time() - iStart))

        # Debugging: mesaure time
        if self.bDoTime: iStart = get_now_time()

        # Figure out which trefwoorden to exclude
        trefwoord_exclude = Trefwoord.objects.filter(toonbaar=0)

        # Create a QSE
        qse = Trefwoord.objects.exclude(id__in=trefwoord_exclude).filter(*lstQ).select_related().order_by(Lower('woord')).distinct()

        # Debugging: time
        if self.bDoTime: 
                print("TrefwoordListView get_queryset part 2: {:.1f}".format(get_now_time() - iStart))
                iStart = get_now_time()

        # Note the number of ITEMS we have
        #   (The nature of these items depends on the approach taken)
        # self.entrycount = qse.count()
        # Using 'len' is faster since [qse] is being actually used again
        self.entrycount = len(qse)

        # Debugging: time
        if self.bDoTime: 
                print("TrefwoordListView get_queryset part 3: {:.1f}".format(get_now_time() - iStart))

        return qse


class LemmaListView(ListView):
    """ListView of lemma's"""

    model = Lemma
    # context_object_name = 'lemma'    
    template_name = 'dictionary/lemma_list.html'
    template_ajax = 'dictionary/lemma_list_oview.html'
    # paginate_by = paginateEntries
    paginate_by = paginateSize
    entrycount = 0
    bUseMijnen = True       # Limburg uses mijnen, Brabant not
    bWbdApproach = True     # Filter using the WBD approach
    bOrderWrdToel = False   # Use the word order 'dialectopgave-toelichting' if True
    bDoTime = True          # Measure time
    qEntry = None
    qs = None
    strict = True      # Use strict filtering ALWAYS

    def get_qs(self):
        """Get the Entry elements that are selected"""
        if self.qEntry == None:
            # Get the [Lemma] pk's that are relevant
            if self.qs != None:
                qs = self.qs
            else:
                # Get the Lemma PKs
                qs = self.get_queryset()
            # Get the Entry queryset related to this
            qs = Entry.objects.filter(lemma__pk__in=qs).select_related()
        else:
            qs = self.qEntry
        return qs

    def render_to_response(self, context, **response_kwargs):
        """Check if a CSV response is needed or not"""
        if 'Csv' in self.request.GET.get('submit_type', ''):
            """ Provide CSV response"""
            
            return export_csv(self.get_qs(), 'begrippen')
        elif 'Excel' in self.request.GET.get('submit_type', ''):
            """ Provide Excel response"""

            return export_xlsx(self.get_qs(), 'begrippen')
        elif 'Html' in self.request.GET.get('submit_type', ''):
            """ Provide Html response"""

            return export_html(self.get_qs(), 'begrippen')

        else:
            iStart = get_now_time()
            # sResp = render_to_string(self.template_name, context)
            oRendered = super(LemmaListView, self).render_to_response(context, **response_kwargs)
            # Show what the render-time was
            print("LemmaListView render time: {:.1f}".format(get_now_time() - iStart))
            return oRendered

    def post(self, request, *args, **kwargs):
        # Prepare return object
        oData = {'status': 'error', 'msg': ''}
        oErr = ErrHandle()
        try:
            self.object_list = self.get_queryset()
            allow_empty = self.get_allow_empty()

            if not allow_empty:
                # When pagination is enabled and object_list is a queryset,
                # it's better to do a cheap query than to load the unpaginated
                # queryset in memory.
                if self.get_paginate_by(self.object_list) is not None and hasattr(self.object_list, 'exists'):
                    is_empty = not self.object_list.exists()
                else:
                    is_empty = len(self.object_list) == 0
                if is_empty:
                    raise Http404(_("Empty list and '%(class_name)s.allow_empty' is False.") % {
                        'class_name': self.__class__.__name__,
                    })
            # Start collecting context time
            if self.bDoTime: iStart = get_now_time()

            context = self.get_context_data()
            if self.bDoTime:
                print("LemmaListView context [a]: {:.1f}".format( get_now_time() - iStart))
                iStart = get_now_time()

            sText = render_to_string(self.template_ajax, context, request)
            if self.bDoTime:
                print("LemmaListView context [b]: {:.1f}".format( get_now_time() - iStart))

            oData['html'] = sText
            oData['status'] = "ok"
        except:
            oData['msg'] = oErr.get_error_message()
        return JsonResponse(oData)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(LemmaListView, self).get_context_data(**kwargs)

        # Get parameters for the search
        # initial = self.request.GET
        initial = self.get
        if initial is None:
            initial = {'optdialect': 'stad'}

        # Possibly set the page object
        if 'page' in initial:
            oPage = context['page_obj']
            oPage.number = int(initial['page'])
            context['page_obj'] = oPage

        # Action depends on the approach
        if self.bWbdApproach:
            # Start collecting context time
            if self.bDoTime: iStart = get_now_time()
            # Need to adapt the object_list to get the entries to be used
            context['object_list'] = list(self.get_entryset(context['page_obj']))
            if self.bDoTime:
                print("LemmaListView get_context_data: {:.1f}".format( get_now_time() - iStart))

        # Start collecting context time
        if self.bDoTime: iStart = get_now_time()

        search_form = LemmaSearchForm(initial)

        context['searchform'] = search_form

        # Get possible user choice of 'strict'
        if 'strict' in initial:
            self.strict = (initial['strict'] == "True")
        context['strict'] = str(self.strict)
        
        # Determine the count 
        context['entrycount'] = self.entrycount #  self.get_queryset().count()

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        if 'paginate_by' in initial:
            context['paginateSize'] = int(initial['paginate_by'])
            self.paginate_by = int(initial['paginate_by'])
        else:
            context['paginateSize'] = self.paginate_by

        if self.bUseMijnen:
            # Try to retain the choice for Mijn
            if 'mijn' in initial:
                mijn_id = int(initial['mijn'])
                context['mijnkeuze'] = mijn_id
                mijn_inst = Mijn.objects.filter(id=mijn_id).first()
                if mijn_inst == None:
                    context['mijnnaam'] = ''
                else:
                    context['mijnnaam'] = mijn_inst.naam
            else:
                context['mijnkeuze'] = 0
                context['mijnnaam'] = ''
        # Process and retain the choice for Aflevering
        if 'aflevering' in initial:
            context['aflkeuze'] = int(initial['aflevering'])
            afl = Aflevering.objects.filter(id=context['aflkeuze']).first()
            if afl == None:
                context['afl'] = ''
            else:
                context['afl'] = afl.get_summary()
        else:
            context['aflkeuze'] = 0
            context['afl'] = ''

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Set the title of the application
        context['title'] = "{} begrippen".format(THIS_DICTIONARY)

        # Set the afleveringen that are available
        context['afleveringen'] = [afl for afl in Aflevering.objects.all()]
        context['mijnen'] = [mijn for mijn in Mijn.objects.all().order_by('naam')]

        # Pass on the word-order boolean
        context['order_word_toel'] = self.bOrderWrdToel

        if self.bDoTime:
            print("LemmaListView context part 1: {:.1f}".format( get_now_time() - iStart))
            # Reset the time
            iStart = get_now_time()

        # If we are in 'strict' mode, we need to deliver the [qlist]
        if self.strict:

            # Transform the paginated queryset into a dict sorted by Dialect/Aflevering
            lAflev = self.get_qafl(context)
            if self.bDoTime:
                print("LemmaListView context get_qafl(): {:.1f}".format( get_now_time() - iStart))
                # Reset the time
                iStart = get_now_time()

            lastDescr = None
            # Get a list with 'first' and 'last' values for each item in the current paginated queryset
            lEntry = self.get_qlist(context)
            if self.bDoTime:
                print("LemmaListView context get_qlist(): {:.1f}".format( get_now_time() - iStart))
                # Reset the time
                iStart = get_now_time()

            # Add the sorted-dialect information to lEntry
            for idx, item in enumerate(lEntry):
                # Start or Finish dialect information
                if item['lemma_gloss']['first']:
                    qsa = []
                    # start a list of all lemma-descriptions taken from the entries belonging to one particular lemma
                    lemma_descr_list = []
                # All: add this entry
                qsa.append(lAflev[idx])
                # If not already present: add description
                iDescrId = item['entry'].descr.id
                if iDescrId not in lemma_descr_list:
                    lemma_descr_list.append(iDescrId)
                if item['lemma_gloss']['last']:
                    # COpy the list of Entry elements sorted by Lemma/Aflevering here
                    lEntry[idx]['alist'] = qsa
                    # OLD: lEntry[idx]['dlist'] = self.get_qdescr(item['entry'])
                    lEntry[idx]['dlist'] = self.get_qdescrlist(lemma_descr_list)
                else:
                    lEntry[idx]['alist'] = None
                    lEntry[idx]['dlist'] = None

            context['qlist'] = lEntry
        # Finish measuring context time
        if self.bDoTime:
            print("LemmaListView context: {:.1f}".format( get_now_time() - iStart))

        # Set the method
        context['method'] = "get"       # Alternative: "ajax"

        # Return the calculated context
        return context

    def get_qlist(self, context):
        """Get a list of Entry elements + first/last information"""

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']
        # Start the output
        html = []
        # Initialize the variables whose changes are important
        if self.bOrderWrdToel:
            lVars = ["lemma_gloss", "trefwoord_woord", "dialectopgave", "toelichting", "dialect_stad"]
            lFuns = [["lemma", "gloss"], ["trefwoord", "woord"], Entry.dialectopgave, Entry.get_toelichting, ["dialect", "stad"]]
        else:
            lVars = ["lemma_gloss", "trefwoord_woord", "toelichting", "dialectopgave", "dialect_stad"]
            lFuns = [["lemma", "gloss"], ["trefwoord", "woord"], Entry.get_toelichting, Entry.dialectopgave, ["dialect", "stad"]]
        # Get a list of items containing 'first' and 'last' information
        lItem = get_item_list(lVars, lFuns, qs)
        # REturn this list
        return lItem

    def get_qafl(self, context):
        """Sort the paginated QS by Lemma/Aflevering into a list"""

        bMethodQset = False

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']

        if bMethodQset:
            # Create a list of entry ids
            id_list = [item.id for item in qs]

            # Create the correctly sorted queryset
            qsd = Entry.objects.filter(Q(id__in=id_list)).select_related().order_by(
                'lemma__gloss', 
                'aflevering__deel__nummer', 
                'aflevering__sectie', 
                'aflevering__aflnum')
        else:
            # qsd = copy.copy(qs)
            # Force evaluation
            qsd = list(qs)
            # Now sort the resulting set
            qsd = sorted(qsd, key=lambda el: (el.lemma.gloss, el.get_aflevering()) )

        # Prepare for processing
        lVarsD = ["lem", "afl"]
        lFunsD = [["lemma", "gloss"], Entry.get_aflevering]
        # Create a list of the items
        lAfl = get_item_list(lVarsD, lFunsD, qsd)
        # Return the result
        return lAfl            

    def get_qdescr(self, entry):
        """OLD: Sort the paginated QS by Lemma/lmdescr.toelichting into a list"""

        # Get the sorted set of [lmdescr] objects for this lemma
        #qsd = []
        ## Walk through the query set
        #for d in entry.lemma.lmdescr.order_by(Lower('toelichting'), Lower('bronnenlijst')): qsd.append(d)
        qsd = [d for d in entry.lemma.lmdescr.order_by(Lower('toelichting'), 'bronnenlijst')]
        # Prepare for processing
        lVarsD = ["descr", "bronnen"]
        lFunsD = [["toelichting"], ["bronnenlijst"]]
        # Create a list of the items
        lDescr = get_item_list(lVarsD, lFunsD, qsd)
        # Return the result
        return lDescr            
      
    def get_qdescrlist(self, descr_id_list):
        """Sort the paginated QS by Lemma/lmdescr.toelichting into a list"""

        # Make a query that gets the indicated id's
        qsd = Description.objects.filter(Q(id__in=descr_id_list)).order_by(Lower('toelichting'), Lower('bronnenlijst'))
        # Prepare for processing
        lVarsD = ["descr", "bronnen"]
        lFunsD = [["toelichting"], ["bronnenlijst"]]
        # Create a list of the items
        lDescr = get_item_list(lVarsD, lFunsD, qsd)
        # Return the result
        return lDescr            
      
    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('paginate_by', self.paginate_by)
        
    def get_entryset(self, page_obj):
        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Initialize timer
        if self.bDoTime: iStart = get_now_time()

        # Retrieve the set of trefwoorden from the page_obj
        lemma_list = [item.id for item in page_obj.object_list]
        if self.bDoTime: print("LemmaListView get_entryset part 1: {:.1f}".format(get_now_time() - iStart))

        if self.bDoTime: iStart = get_now_time()
        # Initialize the filtering
        lstQ.append(Q(lemma__id__in=lemma_list))
        # lstQ.append(Q(lemma__id__in=page_obj))

        # Get the parameters passed on with the GET request
        get = self.get

        # Check for dialect city
        if 'dialectCity' in get and get['dialectCity'] != '':
            val = adapt_search(get['dialectCity'])
            if self.strict:
                lstQ.append(Q(dialect__stad__iregex=val))
            else:
                lstQ.append(Q(entry__dialect__stad__iregex=val))
            bHasFilter = True

        # Check for dialect code (Kloeke)
        if 'dialectCode' in get and get['dialectCode'] != '':
            val = adapt_search(get['dialectCode'])
            if self.strict:
                lstQ.append(Q(dialect__nieuw__iregex=val))
            else:
                lstQ.append(Q(entry__dialect__nieuw__iregex=val))
            bHasFilter = True

        # Check for dialect word, which is a direct member of Entry
        if 'woord' in get and get['woord'] != '':
            val = adapt_search(get['woord'])
            if self.strict:
                lstQ.append(Q(woord__iregex=val))
            else:
                lstQ.append(Q(entry__woord__iregex=val))
            bHasFilter = True

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    if self.strict:
                        lstQ.append(Q(aflevering__id=iVal))
                    else:
                        lstQ.append(Q(entry__aflevering__id=iVal))
                    bHasFilter = True

        # Check for mijn
        if 'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    if self.strict:
                        lstQ.append(Q(mijnlijst__id=iVal))
                    else:
                        lstQ.append(Q(entry__mijnlijst__id=iVal))
                    bHasFilter = True

        # Make sure we filter on aflevering.toonbaar
        if self.strict:
            lstQ.append(Q(aflevering__toonbaar=True))
        else:
            lstQ.append(Q(entry__aflevering__toonbaar=True))

        if self.bDoTime: print("LemmaListView get_entryset part 2: {:.1f}".format(get_now_time() - iStart))

        # Make the QSE available
        # Order: "lemma_gloss", "trefwoord_woord", "dialectopgave", "dialect_stad"
        if self.bDoTime: iStart = get_now_time()
        if self.bOrderWrdToel:
            qse = Entry.objects.filter(*lstQ).distinct().select_related().order_by(
                Lower('lemma__gloss'),  
                Lower('trefwoord__woord'), 
                Lower('woord'), 
                Lower('toelichting'), 
                Lower('dialect__stad'))
        else:
            qse = Entry.objects.filter(*lstQ).distinct().select_related().order_by(
                Lower('lemma__gloss'),  
                Lower('trefwoord__woord'), 
                Lower('toelichting'), 
                Lower('woord'), 
                Lower('dialect__stad'))
        if self.bDoTime: print("LemmaListView get_entryset part 3: {:.1f}".format(get_now_time() - iStart))
        # x = str(Entry.objects.filter(*lstQ).distinct().select_related().query)
        self.qEntry = qse
        return qse

    def get_queryset(self):
        # Measure how long it takes
        if self.bDoTime:
            iStart = get_now_time()

        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        self.get = get

        # Get possible user choice of 'strict'
        if 'strict' in get:
            self.strict = (get['strict'] == "True")

        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Fine-tuning: search string is the LEMMA
        if 'search' in get and get['search'] != '':
            val = get['search']
            if '*' in val or '[' in val or '?' in val:
                val = adapt_search(val)
                lstQ.append(Q(gloss__iregex=val) )
            else:
                # Strive for equality, but disregard case
                lstQ.append(Q(gloss__iexact=val))
            bHasSearch = True

            # check for possible exact numbers having been given
            if re.match('^\d+$', val):
                lstQ.append(Q(sn__exact=val))
 
        # Check for dialect city
        if 'dialectCity' in get and get['dialectCity'] != '':
            val = get['dialectCity']
            if '*' in val or '[' in val or '?' in val:
                # val = adapt_search(get['dialectCity'])
                val = adapt_search(val)
                lstQ.append(Q(entry__dialect__stad__iregex=val))
            else:
                # Strive for equality, but disregard case
                lstQ.append(Q(entry__dialect__stad__iexact=val))
            bHasFilter = True

        # Check for dialect code (Kloeke)
        if 'dialectCode' in get and get['dialectCode'] != '':
            val = get['dialectCode']
            if '*' in val or '[' in val or '?' in val:
                val = adapt_search(val)
                lstQ.append(Q(entry__dialect__nieuw__iregex=val))
            else:
                # Strive for equality, but disregard case
                lstQ.append(Q(entry__dialect__nieuw__iexact=val))
            bHasFilter = True

        # Check for dialect word, which is a direct member of Entry
        if 'woord' in get and get['woord'] != '':
            val = adapt_search(get['woord'])
            lstQ.append(Q(entry__woord__iregex=val))
            bHasFilter = True

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__aflevering__id=iVal))
                    bHasFilter = True

        # Check for mijn
        if self.bUseMijnen and  'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__mijnlijst__id=iVal))
                    bHasFilter = True

        # Method #8 -- use the lemma.toonbaar property
        lemma_hide = Lemma.objects.filter(toonbaar=0)

        qse = Lemma.objects.exclude(id__in=lemma_hide).filter(*lstQ).select_related().order_by('gloss').distinct()

        # Time measurement
        if self.bDoTime:
            print("LemmaListView get_queryset point 'a': {:.1f}".format( get_now_time() - iStart))
            print("LemmaListView query: {}".format(qse.query))

        # Note the number of ITEMS we have
        #   (The nature of these items depends on the approach taken)
        # self.entrycount = qse.count()
        # Note: while taking more time here, it saves time later
        self.entrycount = len(qse)

        # Time measurement
        if self.bDoTime:
            print("LemmaListView get_queryset point 'b': {:.1f}".format( get_now_time() - iStart))

        # Return the resulting filtered and sorted queryset
        return qse


class LocationListView(ListView):
    """Listview of locations"""

    model = Dialect     # The LocationListView uses [Dialect]
    template_name = 'dictionary/location_list.html'
    # paginate_by = paginateEntries # paginateSize
    paginate_by = paginateLocations    # Default pagination number SPECIFICALLY for dialects (1)
    entrycount = 0      # Number of items in queryset (whether Entry or Dialect!!)
    bUseMijnen = True   # Limburg uses mijnen, Brabant not, Gelderland neither
    bWbdApproach = True # Filter using the WBD approach
    bNewOrder = True    # Use the new order (see issue #22 of the RU-wld)
    qEntry = None       # Current queryset as restricted to ENTRY
    qAll = None         # Ordered queryset of ALL
    qs = None           # Current queryset (for speeding up)
    strict = True       # Use strict filtering ALWAYS
    bDoTime = True      # Use timing to determine what goes fastest

    def get_qs(self):
        """Get the Entry elements that are selected"""

        if self.qEntry == None:
            # First get the currently selected elements
            if self.qs != None:
                qs = self.qs
            else:
                # Calculate the  PKs
                qs = self.get_queryset()
            # Get the Entry elements that refer to the set of dialects
            if not self.strict:
                # Convert the [Dialect] elements in qs to [Entry] elements
                qs = Entry.objects.filter(dialect__pk__in=qs).select_related()
        else:
            qs = self.qEntry
        return qs

    def render_to_response(self, context, **response_kwargs):
        """Check if a CSV response is needed or not"""
        if 'Csv' in self.request.GET.get('submit_type', ''):
            """ Provide CSV response"""
            return export_csv(self.get_qs(), 'plaatsen')

        elif 'Excel' in self.request.GET.get('submit_type', ''):
            """ Provide Excel response"""
            return export_xlsx(self.get_qs(), 'plaatsen')
        elif 'Html' in self.request.GET.get('submit_type', ''):
            """ Provide Html response"""
            return export_html(self.get_qs(), 'plaatsen')
        else:
            return super(LocationListView, self).render_to_response(context, **response_kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(LocationListView, self).get_context_data(**kwargs)

        # Start collecting context time
        if self.bDoTime: iStart = get_now_time()

        # Action depends on the approach
        if self.bWbdApproach:
            # Need to adapt the object_list to get the entries to be used
            # context['object_list'] = self.get_entryset(context['page_obj'])
            context['object_list'] = list(self.get_entryset(context['page_obj']))
            if self.bDoTime:
                print("LocationListView context - get_entryset(): {:.1f}".format( get_now_time() - iStart))
                # Reset the time
                iStart = get_now_time()

        # Get parameters for the search
        initial = self.request.GET
        search_form = DialectSearchForm(initial)

        context['searchform'] = search_form

        # Determine the count 
        context['entrycount'] = self.entrycount   #  self.get_queryset().count()

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        if self.bDoTime:
            print("LocationListView context part 0: {:.1f}".format( get_now_time() - iStart))
            # Reset the time
            iStart = get_now_time()

        # Set the afleveringen that are available
        context['afleveringen'] = [afl for afl in Aflevering.objects.all()]

        context['mijnen'] = [mijn for mijn in Mijn.objects.all().order_by('naam')]

        if 'paginate_by' in initial:
            context['paginateSize'] = int(initial['paginate_by'])
            self.paginate_by = int(initial['paginate_by'])
        else:
            context['paginateSize'] = self.paginate_by

        if self.bDoTime:
            print("LocationListView context part 1: {:.1f}".format( get_now_time() - iStart))
            # Reset the time
            iStart = get_now_time()

        # Try to retain the choice for Aflevering and Mijn
        if self.bUseMijnen:
            # Try to retain the choice for Mijn
            if 'mijn' in initial:
                mijn_id = int(initial['mijn'])
                context['mijnkeuze'] = mijn_id
                mijn_inst = Mijn.objects.filter(id=mijn_id).first()
                if mijn_inst == None:
                    context['mijnnaam'] = ''
                else:
                    context['mijnnaam'] = mijn_inst.naam
            else:
                context['mijnkeuze'] = 0
                context['mijnnaam'] = ''

        if 'aflevering' in initial:
            context['aflkeuze'] = int(initial['aflevering'])
            afl = Aflevering.objects.filter(id=context['aflkeuze']).first()
            if afl == None:
                context['afl'] = ''
            else:
                context['afl'] = afl.get_summary()
        else:
            context['aflkeuze'] = 0
            context['afl'] = ''

        if self.bDoTime:
            print("LocationListView context part 2: {:.1f}".format( get_now_time() - iStart))
            # Reset the time
            iStart = get_now_time()

        # Set the title of the application
        context['title'] = "{} plaatsen".format(THIS_DICTIONARY)

        # Get possible user choice of 'strict'
        if 'strict' in initial:
            self.strict = (initial['strict'] == "True")
        context['strict'] = str(self.strict)

        if self.strict:
            # Transform the paginated queryset into a dict sorted by Dialect/Aflevering
            lDialect = self.get_qafl(context)
            if self.bDoTime:
                print("LocationListView context get_qafl(): {:.1f}".format( get_now_time() - iStart))
                # Reset the time
                iStart = get_now_time()

            # Get a list with 'first' and 'last' values for each item in the current paginated queryset
            lEntry = self.get_qlist(context)

            if self.bDoTime:
                print("LocationListView context get_qlist(): {:.1f}".format( get_now_time() - iStart))
                # Reset the time
                iStart = get_now_time()

            # Add the sorted-dialect information to lEntry
            for idx, item in enumerate(lEntry):
                # Start or Finish dialect information
                if item['dialect_stad']['first']:
                    iDialectStart = idx
                    qsa = []
                # All: add this entry
                qsa.append(lDialect[idx])
                if item['dialect_stad']['last']:
                    # COpy the list of Entry elements sorted by Stad/Aflevering here
                    lEntry[idx]['alist'] = qsa
                else:
                    lEntry[idx]['alist'] = None

            context['qlist'] = lEntry

        # Finish measuring context time
        if self.bDoTime:
            print("LocationListView context: {:.1f}".format( get_now_time() - iStart))

        # Return the calculated context
        return context
      
    def get_qlist(self, context):
        """Calculate HTML output for the query-set in the context"""

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']
        # Start the output
        html = []
        # Initialize the variables whose changes are important
        lVars = ["dialect_stad", "lemma_gloss", "trefwoord_woord","toelichting", "dialectopgave"]
        lFuns = [["dialect", "stad"], ["lemma", "gloss"], ["trefwoord", "woord"], Entry.get_toelichting, Entry.dialectopgave]
        # Get a list of items containing 'first' and 'last' information
        lItem = get_item_list(lVars, lFuns, qs)
        # REturn this list
        return lItem

    def get_qafl(self, context):
        """Sort the paginated QS by Dialect/Aflevering and turn into a dict"""

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']
        qsd = list(qs)
        # Now sort the resulting set
        qsd = sorted(qsd, key=lambda el: el.dialect.stad + " " + el.get_aflevering())

        # Prepare for dialect processing
        lVarsD = ["stad", "afl"]
        lFunsD = [["dialect", "stad"], Entry.get_aflevering]
        # Create a list of Dialect items
        lDialect = get_item_list(lVarsD, lFunsD, qsd)
        # Return the result
        return lDialect            

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('paginate_by', self.paginate_by)
        
    def get_entryset(self, page_obj):
        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Time measurement
        if self.bDoTime: iStart = get_now_time()

        # Initialize the filtering on ENTRY
        dialect_list = [item.id for item in page_obj]
        lstQ.append(Q(dialect__id__in=dialect_list))

        # Make sure we filter on aflevering.toonbaar
        lstQ.append(Q(aflevering__toonbaar=True))

        # Time measurement
        if self.bDoTime:
            print("LocationListView get_entryset (point 'e:a'): {:.1f}".format( get_now_time() - iStart))
            iStart = get_now_time()

        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(aflevering__id=iVal) )
                    bHasFilter = True

        # Check for mijn
        if self.bUseMijnen and 'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(mijnlijst__id=iVal) )
                    bHasFilter = True

        bUseLower = True
        if bUseLower:
            qse = Entry.objects.filter(*lstQ).distinct().select_related().order_by(
                Lower('dialect__stad'),
                'lemma__gloss',  
                Lower('trefwoord__woord'), 
                Lower('toelichting'), 
                Lower('woord'))
        else:
            qse = Entry.objects.filter(*lstQ).distinct().select_related().order_by(
                'dialect__stad',
                'lemma__gloss',  
                'trefwoord__woord', 
                'toelichting', 
                'woord')

        # Time measurement
        if self.bDoTime:
            print("LocationListView get_entryset (point 'e:b'): {:.1f}".format( get_now_time() - iStart))

        self.qEntry = qse

        qse = list(qse)
        # Time measurement
        if self.bDoTime:
            print("LocationListView get_entryset (point 'e:c'): {:.1f}".format( get_now_time() - iStart))

        return qse

    def get_queryset(self):
        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        # Not sure why, but get a copy
        get = get.copy()

        # Measure how long it takes
        if self.bDoTime:
            iStart = get_now_time()

        # Set the [sortOrder] parameter to 'stad' (the name of the city)
        get['sortOrder'] = 'stad'

        # Get possible user choice of 'strict'
        if 'strict' in get:
            self.strict = (get['strict'] == "True")

        # Queryset: build a list of requirements
        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Fine-tuning: search string is the STAD
        if 'search' in get and get['search'] != '':
            val = adapt_search(get['search'])
            lstQ.append(Q(stad__iregex=val) )
            bHasSearch = True

            # check for possible exact numbers having been given
            if re.match('^\d+$', val):
                query = query | Q(sn__exact=val)
                lstQ.append(Q(sn__exact=val))

        # Check for dialect code (Kloeke)
        if 'nieuw' in get and get['nieuw'] != '':
            val = adapt_search(get['nieuw'])
            lstQ.append(Q(nieuw__iregex=val) )
            bHasSearch = True

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__aflevering__id=iVal) )
                    bHasFilter = True

        # Check for mijn
        if self.bUseMijnen and 'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__mijnlijst__id=iVal) )
                    bHasFilter = True

        # Time measurement
        if self.bDoTime:
            print("LocationListView get_queryset point 'a': {:.1f}".format( get_now_time() - iStart))
            iStart = get_now_time()

        # Get a list of Dialects that should be excluded
        dialect_hide = Dialect.objects.filter(toonbaar=0)

        # Use the E-WBD approach: be efficient here
        qs = Dialect.objects.exclude(id__in=dialect_hide).filter(*lstQ).distinct().select_related().order_by(Lower('stad'))

        # Time measurement
        if self.bDoTime:
            print("LocationListView get_queryset point 'c': {:.1f}".format( get_now_time() - iStart))
            iStart = get_now_time()

        # self.entrycount = qs.count()
        # Using 'len' is faster since [qse] is being actually used again
        self.entrycount = len(qs)

        # Time measurement
        if self.bDoTime:
            print("LocationListView get_queryset point 'd': {:.1f}".format( get_now_time() - iStart))
            iStart = get_now_time()

        # Return the resulting filtered and sorted queryset
        return qs


class DialectListView(ListView):
    """Listview of dialects"""

    model = Dialect
    paginate_by = 10
    template_name = 'dictionary/dialect_list.html'
    entrycount = 0
    bDoTime = True

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(DialectListView, self).get_context_data(**kwargs)

        # Get parameters for the search
        initial = self.request.GET
        search_form = DialectSearchForm(initial)

        context['searchform'] = search_form

        # Determine the count 
        context['entrycount'] = self.get_queryset().count()

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        if 'paginate_by' in initial:
            context['paginateSize'] = int(initial['paginate_by'])
        else:
            context['paginateSize'] = paginateSize

        # Set the title of the application
        context['title'] = "{} dialecten".format(THIS_DICTIONARY)

        # Return the calculated context
        return context

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('paginate_by', self.paginate_by)
        
    def get_queryset(self):
        # Measure how long it takes
        if self.bDoTime: iStart = get_now_time()

        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        get = get.copy()
        self.get = get

        # Fix the sort-order
        get['sortOrder'] = 'stad'

        lstQ = []

        # Fine-tuning: search string is the LEMMA
        if 'search' in get and get['search'] != '':
            val = adapt_search(get['search'])
            query = Q(stad__iregex=val) 

            # check for possible exact numbers having been given
            if re.match('^\d+$', val):
                query = query | Q(sn__exact=val)

            # Apply the filter
            lstQ.append(query)

        # Check for dialect code (Kloeke)
        if 'nieuw' in get and get['nieuw'] != '':
            val = adapt_search(get['nieuw'])
            query = Q(nieuw__iregex=val)
            
            # Apply the filter
            lstQ.append(query)

        # Calculate the final qs
        qs = Dialect.objects.exclude(toonbaar=0).filter(*lstQ).order_by('stad').distinct()

        # Time measurement
        if self.bDoTime:
            print("DialectListView get_queryset point 'a': {:.1f}".format( get_now_time() - iStart))
            print("DialectListView query: {}".format(qs.query))
            iStart = get_now_time()

        # Determine the length
        self.entrycount = len(qs)

        # Time measurement
        if self.bDoTime:
            print("DialectListView get_queryset point 'b': {:.1f}".format( get_now_time() - iStart))

        # Return the resulting filtered and sorted queryset
        return qs


class MijnListView(ListView):
    """Listview of mines"""

    model = Mijn
    paginate_by = 10
    template_name = 'dictionary/mijn_list.html'


    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(MijnListView, self).get_context_data(**kwargs)

        # Get parameters for the search
        initial = self.request.GET
        search_form = MijnSearchForm(initial)

        context['searchform'] = search_form

        # Determine the count 
        context['entrycount'] = self.get_queryset().count()

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        if 'paginate_by' in initial:
            context['paginateSize'] = int(initial['paginate_by'])
        else:
            context['paginateSize'] = paginateSize

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Set the title of the application
        context['title'] = "{} mijnen".format(THIS_DICTIONARY)

        # Return the calculated context
        return context

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('paginate_by', self.paginate_by)
        
    def get_queryset(self):

        # Get the parameters passed on with the GET request
        get = self.request.GET.copy()
        get['sortOrder'] = 'naam'

        # Queryset: start out with *ALL* the mines
        qs = Mijn.objects.all()

        # Fine-tuning: search string is the LEMMA
        if 'search' in get and get['search'] != '':
            val = adapt_search(get['search'])
            # The main search is on the NAME of the mine
            query = Q(naam__iregex=val) 

            # check for possible exact numbers having been given
            if re.match('^\d+$', val):
                query = query | Q(sn__exact=val)

            # Apply the filter
            qs = qs.filter(query)

        # Check for toelichting
        if 'toelichting' in get and get['toelichting'] != '':
            val = adapt_search(get['toelichting'])
            # query = Q(nieuw__istartswith=val)
            query = Q(toelichting__iregex=val)
            qs = qs.filter(query)

        # Check for locatie
        if 'locatie' in get and get['locatie'] != '':
            val = adapt_search(get['locatie'])
            # query = Q(nieuw__istartswith=val)
            query = Q(locatie__iregex=val)
            qs = qs.filter(query)

        # Make sure we only have distinct values
        qs = qs.distinct()

        # Sort the queryset by the parameters given
        qs = order_queryset_by_sort_order(get, qs, 'naam')

        # Return the resulting filtered and sorted queryset
        return qs


class DeelListView(ListView):
    """Overview of all the [aflevering] books per [deel]"""

    model = Deel
    template_name = 'dictionary/aflevering_list.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(DeelListView, self).get_context_data(**kwargs)

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Set the title of the application
        context['title'] = "{} afleveringen".format(THIS_DICTIONARY)

        context['intro_pdf'] = ""
        context['intro_op_drie_pdf'] = ""

        # Return the calculated context
        return context

